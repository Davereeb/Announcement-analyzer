import sqlite3
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter

import config

COLUMNS = [
    "股票代码", "股票名称", "发布时间", "公告标题", "公告类型", "公告链接",
    "is_valuable", "summary", "reason", "emotion",
    "doc_type", "text_length", "fetch_time", "llm_time",
    "input_tokens", "output_tokens", "processed_at",
]


def _extract_stock_name(title: str) -> str:
    """从'股票名称:公告内容'格式的标题中提取股票名称。"""
    if title and ":" in title:
        return title.split(":")[0].strip()
    if title and "：" in title:
        return title.split("：")[0].strip()
    return ""

EMOTION_FILLS = {
    2:  PatternFill(fill_type="solid", fgColor="00B050"),   # green
    1:  PatternFill(fill_type="solid", fgColor="92D050"),   # light green
    0:  PatternFill(fill_type="solid", fgColor="FFFFFF"),   # white
    -1: PatternFill(fill_type="solid", fgColor="FFCCCC"),   # light red
    -2: PatternFill(fill_type="solid", fgColor="FF0000"),   # red
}


def export_to_excel():
    conn = sqlite3.connect(config.OUTPUT_DB)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT * FROM announcements WHERE status='done'"
    ).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "results"

    # Header row
    ws.append(COLUMNS)
    ws.freeze_panes = "A2"

    emotion_col_idx = COLUMNS.index("emotion") + 1
    summary_col_idx = COLUMNS.index("summary") + 1

    for row_data in rows:
        row = []
        for col in COLUMNS:
            if col == "股票名称":
                row.append(_extract_stock_name(row_data["公告标题"] or ""))
            else:
                row.append(row_data[col] if col in row_data.keys() else None)
        ws.append(row)

        excel_row = ws.max_row
        emotion_val = row_data["emotion"]
        if emotion_val is not None and emotion_val in EMOTION_FILLS:
            cell = ws.cell(row=excel_row, column=emotion_col_idx)
            cell.fill = EMOTION_FILLS[emotion_val]

        # Wrap summary cell
        summary_cell = ws.cell(row=excel_row, column=summary_col_idx)
        summary_cell.alignment = Alignment(wrap_text=True)

    # Column widths
    ws.column_dimensions[get_column_letter(summary_col_idx)].width = 60
    # Reasonable widths for other columns
    default_widths = {
        "股票代码": 12, "股票名称": 14, "发布时间": 14, "公告标题": 40, "公告类型": 16,
        "公告链接": 20, "is_valuable": 10, "reason": 30, "emotion": 8,
        "doc_type": 10, "text_length": 12, "fetch_time": 12, "llm_time": 10,
        "input_tokens": 14, "output_tokens": 14, "processed_at": 20,
    }
    for col_name, width in default_widths.items():
        if col_name in COLUMNS:
            idx = COLUMNS.index(col_name) + 1
            ws.column_dimensions[get_column_letter(idx)].width = width

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"results_{timestamp}.xlsx"
    wb.save(filename)
    print(f"已导出 {len(rows)} 条记录 → {filename}")
    return filename
